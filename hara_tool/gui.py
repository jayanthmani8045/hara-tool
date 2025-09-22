"""
GUI module for HARA Tool - Main window and dialogs
"""

import os
from pathlib import Path
import pandas as pd
from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QComboBox, QTableWidget,
    QTableWidgetItem, QTextEdit, QGroupBox, QProgressBar,
    QMessageBox, QHeaderView, QTabWidget, QFrame,
    QCheckBox, QSlider, QDialog, QSpinBox, QFormLayout, QDialogButtonBox
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont, QColor, QIcon

from .processor import ExcelProcessor
from .constants import ASIL_COLORS, DEFAULT_SETTINGS
from . import styles


class AdvancedMatchingDialog(QDialog):
    """Dialog for advanced matching settings"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Advanced Matching Settings")
        self.setModal(True)
        self.setStyleSheet(parent.styleSheet() if parent else "")
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Case sensitivity
        self.case_sensitive = QCheckBox("Case Sensitive Matching")
        layout.addWidget(self.case_sensitive)
        
        # Strip whitespace
        self.strip_whitespace = QCheckBox("Strip Extra Whitespace")
        self.strip_whitespace.setChecked(True)
        layout.addWidget(self.strip_whitespace)
        
        # Weight settings for combined matching
        weight_group = QGroupBox("Match Weights")
        weight_layout = QFormLayout()
        
        self.os_weight = QSpinBox()
        self.os_weight.setRange(0, 100)
        self.os_weight.setValue(70)
        self.os_weight.setSuffix("%")
        weight_layout.addRow("Operating Scenario Weight:", self.os_weight)
        
        self.hazard_weight = QSpinBox()
        self.hazard_weight.setRange(0, 100)
        self.hazard_weight.setValue(30)
        self.hazard_weight.setSuffix("%")
        weight_layout.addRow("Hazard Weight:", self.hazard_weight)
        
        weight_group.setLayout(weight_layout)
        layout.addWidget(weight_group)
        
        # Buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | 
            QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
        self.setLayout(layout)


class HARAMainWindow(QMainWindow):
    """Main application window"""
    
    def __init__(self):
        super().__init__()
        self.user_file = None
        self.result_df = None
        self.advanced_settings = {
            'case_sensitive': DEFAULT_SETTINGS['case_sensitive'],
            'strip_whitespace': DEFAULT_SETTINGS['strip_whitespace'],
            'os_weight': DEFAULT_SETTINGS['os_weight'],
            'hazard_weight': DEFAULT_SETTINGS['hazard_weight']
        }
        
        self.set_window_icon()
        self.init_ui()
    
    def set_window_icon(self):
        """Set application icon"""
        icon_path = Path(__file__).parent.parent / 'resources' / 'icon.ico'
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))
    
    def init_ui(self):
        """Initialize the user interface"""
        self.setWindowTitle("HARA Automation Tool - Professional Edition")
        self.setGeometry(100, 100, 1500, 900)
        
        # Apply theme
        self.setStyleSheet(styles.DARK_THEME)
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Title Section
        self.create_title_section(main_layout)
        
        # File Selection Section
        self.create_file_section(main_layout)
        
        # Processing Section
        self.create_processing_section(main_layout)
        
        # Results Section
        self.create_results_section(main_layout)
        
        # Save Section
        self.create_save_section(main_layout)
        
        # Initial log message
        self.log("System initialized. ASIL determination table embedded.")
        self.log("Fuzzy matching enabled with 80% threshold (Ratio algorithm).")
    
    def create_title_section(self, parent_layout):
        """Create title section"""
        title_frame = QFrame()
        title_layout = QVBoxLayout(title_frame)
        
        title_label = QLabel("HARA AUTOMATION TOOL")
        title_label.setStyleSheet(styles.TITLE_STYLE)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_layout.addWidget(title_label)
        
        subtitle_label = QLabel("Automated Safety Integrity Level Determination")
        subtitle_label.setStyleSheet(styles.SUBTITLE_STYLE)
        subtitle_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_layout.addWidget(subtitle_label)
        
        separator = QFrame()
        separator.setObjectName("separator")
        separator.setFixedHeight(1)
        title_layout.addWidget(separator)
        
        parent_layout.addWidget(title_frame)
    
    def create_file_section(self, parent_layout):
        """Create file selection section"""
        file_group = QGroupBox("FILE SELECTION")
        file_layout = QVBoxLayout()
        
        # User file selection
        user_file_layout = QHBoxLayout()
        file_label_title = QLabel("Excel File:")
        file_label_title.setFixedWidth(100)
        user_file_layout.addWidget(file_label_title)
        
        self.file_label = QLabel("No file selected")
        self.file_label.setStyleSheet(styles.FILE_LABEL_STYLE_DEFAULT)
        user_file_layout.addWidget(self.file_label, 1)
        
        self.browse_btn = QPushButton("Browse")
        self.browse_btn.clicked.connect(self.browse_file)
        user_file_layout.addWidget(self.browse_btn)
        file_layout.addLayout(user_file_layout)
        
        # Sheet selection
        sheet_layout = QHBoxLayout()
        
        os_label = QLabel("Operating Scenario Sheet:")
        os_label.setFixedWidth(180)
        sheet_layout.addWidget(os_label)
        self.os_combo = QComboBox()
        self.os_combo.setEnabled(False)
        sheet_layout.addWidget(self.os_combo, 1)
        
        sheet_layout.addSpacing(30)
        
        ra_label = QLabel("Risk Assessment Sheet:")
        ra_label.setFixedWidth(180)
        sheet_layout.addWidget(ra_label)
        self.ra_combo = QComboBox()
        self.ra_combo.setEnabled(False)
        sheet_layout.addWidget(self.ra_combo, 1)
        
        file_layout.addLayout(sheet_layout)
        file_group.setLayout(file_layout)
        parent_layout.addWidget(file_group)
    
    def create_processing_section(self, parent_layout):
        """Create processing section with fuzzy matching controls"""
        process_group = QGroupBox("PROCESSING & MATCHING SETTINGS")
        process_layout = QVBoxLayout()
        
        # Fuzzy Matching Controls
        fuzzy_layout = QHBoxLayout()
        
        self.fuzzy_enabled = QCheckBox("Enable Fuzzy Matching")
        self.fuzzy_enabled.setChecked(True)
        self.fuzzy_enabled.setStyleSheet(styles.FUZZY_CHECKBOX_STYLE)
        self.fuzzy_enabled.toggled.connect(self.on_fuzzy_toggle)
        fuzzy_layout.addWidget(self.fuzzy_enabled)
        
        fuzzy_layout.addWidget(QLabel("Threshold:"))
        self.fuzzy_slider = QSlider(Qt.Orientation.Horizontal)
        self.fuzzy_slider.setMinimum(50)
        self.fuzzy_slider.setMaximum(100)
        self.fuzzy_slider.setValue(80)
        self.fuzzy_slider.setTickPosition(QSlider.TickPosition.TicksBelow)
        self.fuzzy_slider.setTickInterval(10)
        self.fuzzy_slider.setFixedWidth(200)
        self.fuzzy_slider.valueChanged.connect(self.on_threshold_change)
        fuzzy_layout.addWidget(self.fuzzy_slider)
        
        self.threshold_label = QLabel("80%")
        self.threshold_label.setStyleSheet(styles.THRESHOLD_LABEL_STYLE)
        fuzzy_layout.addWidget(self.threshold_label)
        
        fuzzy_layout.addWidget(QLabel("Algorithm:"))
        self.algorithm_combo = QComboBox()
        self.algorithm_combo.addItems([
            "Ratio (Default)",
            "Partial Ratio",
            "Token Sort Ratio",
            "Token Set Ratio"
        ])
        self.algorithm_combo.setToolTip(
            "Ratio: Simple string similarity\n"
            "Partial: Matches substring\n"
            "Token Sort: Ignores word order\n"
            "Token Set: Ignores duplicates and order"
        )
        fuzzy_layout.addWidget(self.algorithm_combo)
        
        self.advanced_btn = QPushButton("Advanced")
        self.advanced_btn.clicked.connect(self.show_advanced_settings)
        fuzzy_layout.addWidget(self.advanced_btn)
        
        fuzzy_layout.addStretch()
        process_layout.addLayout(fuzzy_layout)
        
        # Process button and progress
        process_btn_layout = QHBoxLayout()
        self.process_btn = QPushButton("Process HARA")
        self.process_btn.setEnabled(False)
        self.process_btn.setStyleSheet(styles.PROCESS_BUTTON_STYLE)
        self.process_btn.clicked.connect(self.process_data)
        process_btn_layout.addWidget(self.process_btn)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(True)
        process_btn_layout.addWidget(self.progress_bar, 1)
        
        self.status_label = QLabel("")
        self.status_label.setStyleSheet(styles.STATUS_LABEL_STYLE)
        process_btn_layout.addWidget(self.status_label)
        
        process_layout.addLayout(process_btn_layout)
        process_group.setLayout(process_layout)
        parent_layout.addWidget(process_group)
    
    def create_results_section(self, parent_layout):
        """Create results section with tabs"""
        self.tabs = QTabWidget()
        
        # Results Table Tab
        self.result_table = QTableWidget()
        self.result_table.setAlternatingRowColors(True)
        self.tabs.addTab(self.result_table, "Results")
        
        # Log Tab
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.tabs.addTab(self.log_text, "Processing Log")
        
        parent_layout.addWidget(self.tabs, 1)
    
    def create_save_section(self, parent_layout):
        """Create save/export section"""
        save_group = QGroupBox("EXPORT OPTIONS")
        save_layout = QHBoxLayout()
        
        self.save_btn = QPushButton("Save to Excel")
        self.save_btn.setToolTip("Add RESULT sheet to original file")
        self.save_btn.setEnabled(False)
        self.save_btn.clicked.connect(self.save_results)
        save_layout.addWidget(self.save_btn)
        
        self.export_btn = QPushButton("Export New File")
        self.export_btn.setToolTip("Export results to a new Excel file")
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_results)
        save_layout.addWidget(self.export_btn)
        
        save_layout.addStretch()
        
        info_label = QLabel("ASIL Determination: Embedded | Valid Ranges: S(0-3) C(0-3) E(1-4)")
        info_label.setStyleSheet(styles.INFO_LABEL_STYLE)
        save_layout.addWidget(info_label)
        
        save_group.setLayout(save_layout)
        parent_layout.addWidget(save_group)
    
    def on_fuzzy_toggle(self, checked):
        """Handle fuzzy matching toggle"""
        self.fuzzy_slider.setEnabled(checked)
        self.algorithm_combo.setEnabled(checked)
        if checked:
            self.log(f"Fuzzy matching enabled (threshold: {self.fuzzy_slider.value()}%)")
        else:
            self.log("Fuzzy matching disabled - exact matches only")
    
    def on_threshold_change(self, value):
        """Handle threshold slider change"""
        self.threshold_label.setText(f"{value}%")
    
    def show_advanced_settings(self):
        """Show advanced matching settings dialog"""
        dialog = AdvancedMatchingDialog(self)
        
        # Load current settings
        dialog.case_sensitive.setChecked(self.advanced_settings['case_sensitive'])
        dialog.strip_whitespace.setChecked(self.advanced_settings['strip_whitespace'])
        dialog.os_weight.setValue(self.advanced_settings['os_weight'])
        dialog.hazard_weight.setValue(self.advanced_settings['hazard_weight'])
        
        if dialog.exec():
            # Save settings
            self.advanced_settings['case_sensitive'] = dialog.case_sensitive.isChecked()
            self.advanced_settings['strip_whitespace'] = dialog.strip_whitespace.isChecked()
            self.advanced_settings['os_weight'] = dialog.os_weight.value()
            self.advanced_settings['hazard_weight'] = dialog.hazard_weight.value()
            
            self.log(f"Advanced settings updated: Case sensitive={self.advanced_settings['case_sensitive']}, "
                    f"Strip whitespace={self.advanced_settings['strip_whitespace']}, "
                    f"Weights: OS={self.advanced_settings['os_weight']}%, "
                    f"Hazard={self.advanced_settings['hazard_weight']}%")
    
    def browse_file(self):
        """Browse for user's Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            self.user_file = file_path
            self.file_label.setText(Path(file_path).name)
            self.file_label.setStyleSheet(styles.FILE_LABEL_STYLE_SELECTED)
            self.load_sheets()
    
    def load_sheets(self):
        """Load sheet names from Excel file"""
        try:
            xl_file = pd.ExcelFile(str(self.user_file))
            sheets = xl_file.sheet_names
            
            # Filter out HAZOP sheet if it exists
            filtered_sheets = [sheet for sheet in sheets if 'hazop' not in sheet.lower()]
            
            self.os_combo.clear()
            self.os_combo.addItems(filtered_sheets)
            self.os_combo.setEnabled(True)
            
            self.ra_combo.clear()
            self.ra_combo.addItems(filtered_sheets)
            self.ra_combo.setEnabled(True)
            
            # Try to auto-select sheets based on names
            for i, sheet in enumerate(filtered_sheets):
                sheet_lower = sheet.lower()
                if 'operating' in sheet_lower and 'scenario' in sheet_lower:
                    self.os_combo.setCurrentIndex(i)
                elif 'risk' in sheet_lower and 'assessment' in sheet_lower:
                    self.ra_combo.setCurrentIndex(i)
            
            self.process_btn.setEnabled(True)
            self.log(f"File loaded: {len(filtered_sheets)} sheets available")
            self.status_label.setText("Ready")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load Excel file: {str(e)}")
            self.log(f"ERROR: Failed to load file - {str(e)}")
    
    def process_data(self):
        """Start processing in background thread"""
        if not self.user_file:
            QMessageBox.warning(self, "Warning", "Please select an Excel file first")
            return
        
        os_sheet = self.os_combo.currentText()
        ra_sheet = self.ra_combo.currentText()
        
        if os_sheet == ra_sheet:
            QMessageBox.warning(self, "Warning", 
                              "Please select different sheets for Operating Scenario and Risk Assessment")
            return
        
        # Prepare fuzzy settings
        fuzzy_settings = {
            'enabled': self.fuzzy_enabled.isChecked(),
            'threshold': self.fuzzy_slider.value(),
            'algorithm': self.algorithm_combo.currentText(),
            'case_sensitive': self.advanced_settings['case_sensitive'],
            'strip_whitespace': self.advanced_settings['strip_whitespace'],
            'os_weight': self.advanced_settings['os_weight'],
            'hazard_weight': self.advanced_settings['hazard_weight']
        }
        
        # Disable controls during processing
        self.process_btn.setEnabled(False)
        self.save_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("Processing...")
        
        # Clear previous results
        self.result_table.clear()
        self.tabs.setCurrentIndex(1)  # Switch to log tab
        
        # Create and start processor thread
        self.processor = ExcelProcessor(
            self.user_file,
            os_sheet,
            ra_sheet,
            fuzzy_settings
        )
        
        self.processor.progress.connect(self.update_progress)
        self.processor.log.connect(self.log)
        self.processor.finished.connect(self.on_process_complete)
        self.processor.error.connect(self.on_process_error)
        
        self.log(f"Processing started: {os_sheet} -> {ra_sheet}")
        self.processor.start()
    
    def update_progress(self, value):
        """Update progress bar"""
        self.progress_bar.setValue(value)
    
    def log(self, message):
        """Add message to log with timestamp"""
        timestamp = pd.Timestamp.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{timestamp}] {message}")
    
    def on_process_complete(self, result_df):
        """Handle successful processing"""
        self.result_df = result_df
        
        # Display results in table
        self.display_results(result_df)
        
        # Enable save buttons
        self.save_btn.setEnabled(True)
        self.export_btn.setEnabled(True)
        
        # Re-enable controls
        self.process_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        
        # Count ASIL statistics
        asil_count = len(result_df[result_df['ASIL'] != ''])
        total_count = len(result_df)
        
        # Count S and C values
        s_count = len(result_df[(result_df['S'] != '') & (pd.notna(result_df['S']))])
        c_count = len(result_df[(result_df['C'] != '') & (pd.notna(result_df['C']))])
        
        self.status_label.setText(f"Complete: {total_count} rows processed, {asil_count} with ASIL")
        
        # Switch to results tab
        self.tabs.setCurrentIndex(0)
        
        # Show summary with detailed information
        if asil_count > 0:
            asil_summary = result_df[result_df['ASIL'] != '']['ASIL'].value_counts().to_dict()
            summary_text = ', '.join([f"{k}: {v}" for k, v in sorted(asil_summary.items())])
            message = (f"Processing completed successfully.\n\n"
                      f"Total Scenarios: {total_count}\n"
                      f"ASIL Determined: {asil_count}\n"
                      f"S values found: {s_count}\n"
                      f"C values found: {c_count}\n\n"
                      f"Distribution:\n{summary_text}")
        else:
            # Provide detailed troubleshooting information
            message = (f"Processing completed.\n\n"
                      f"Total Scenarios: {total_count}\n"
                      f"No ASIL values determined\n\n"
                      f"S (Severity) values found: {s_count}\n"
                      f"C (Controllability) values found: {c_count}\n\n")
            
            if s_count == 0 and c_count == 0:
                message += ("Troubleshooting:\n"
                          "• S and C columns not found or empty in Risk Assessment sheet\n"
                          "• Ensure your Risk Assessment sheet has columns named:\n"
                          "  - 'S' or 'Severity' with values 0-3\n"
                          "  - 'C' or 'Controllability' with values 0-3\n"
                          "• Check that matched scenarios have S and C values populated")
            elif s_count == 0:
                message += ("Troubleshooting:\n"
                          "• S (Severity) values not found\n"
                          "• Check that the S column has valid values (0-3)")
            elif c_count == 0:
                message += ("Troubleshooting:\n"
                          "• C (Controllability) values not found\n"
                          "• Check that the C column has valid values (0-3)")
            else:
                message += ("Troubleshooting:\n"
                          "• Some S/C values found but not for matched scenarios\n"
                          "• Check the fuzzy matching threshold\n"
                          "• Verify Operating Scenario names match between sheets")
            
            message += "\n\nPlease check the Processing Log tab for detailed information."
        
        QMessageBox.information(self, "Processing Complete", message)
    
    def on_process_error(self, error_msg):
        """Handle processing error"""
        self.log(f"ERROR: {error_msg}")
        self.status_label.setText("Error")
        QMessageBox.critical(self, "Processing Error", error_msg)
        
        # Re-enable controls
        self.process_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
    
    def display_results(self, df):
        """Display DataFrame in table widget with color coding"""
        self.result_table.setRowCount(len(df))
        self.result_table.setColumnCount(len(df.columns))
        self.result_table.setHorizontalHeaderLabels(df.columns.tolist())
        
        for row in range(len(df)):
            for col in range(len(df.columns)):
                value = df.iloc[row, col]
                if pd.isna(value):
                    value = ""
                item = QTableWidgetItem(str(value))
                
                # Color code ASIL column
                if df.columns[col] == 'ASIL' and value in ASIL_COLORS:
                    color = ASIL_COLORS[value]
                    item.setBackground(QColor(*color))
                    item.setForeground(QColor(255, 255, 255))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                
                # Highlight match type column
                if df.columns[col] == 'Match_Type':
                    if 'Exact' in str(value):
                        item.setForeground(QColor(46, 125, 50))  # Green for exact
                    elif 'Fuzzy' in str(value):
                        item.setForeground(QColor(94, 179, 246))  # Light blue for fuzzy
                    else:
                        item.setForeground(QColor(255, 152, 0))  # Orange for no match
                
                self.result_table.setItem(row, col, item)
        
        # Adjust column widths
        self.result_table.resizeColumnsToContents()
        header = self.result_table.horizontalHeader()
        header.setStretchLastSection(True)
    
    def save_results(self):
        """Save results to RESULT sheet in original file"""
        if self.result_df is None:
            return
        
        try:
            # Read existing Excel file
            with pd.ExcelFile(str(self.user_file)) as xl:
                sheets = {sheet: xl.parse(sheet) for sheet in xl.sheet_names}
            
            # Add or replace RESULT sheet
            sheets['RESULT'] = self.result_df
            
            # Write back to file
            with pd.ExcelWriter(str(self.user_file), engine='openpyxl') as writer:
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, sheet_name=str(sheet_name), index=False)
            
            self.log(f"Results saved to RESULT sheet")
            self.status_label.setText("Saved")
            QMessageBox.information(self, "Success", "Results saved to RESULT sheet successfully.")
            
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save results: {str(e)}")
            self.log(f"ERROR: Save failed - {str(e)}")
    
    def export_results(self):
        """Export results to a new Excel file"""
        if self.result_df is None:
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Results",
            "HARA_Results.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                # Export with formatting
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    self.result_df.to_excel(writer, sheet_name='RESULT', index=False)
                    
                    # Get the workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['RESULT']
                    
                    # Apply professional formatting
                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                    
                    # Header formatting
                    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", 
                                            fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                    
                    # ASIL column formatting
                    asil_col = None
                    for idx, col in enumerate(self.result_df.columns, 1):
                        if col == 'ASIL':
                            asil_col = idx
                            break
                    
                    if asil_col:
                        asil_fills = {
                            'QM': PatternFill(start_color="2E7D32", end_color="2E7D32", 
                                            fill_type="solid"),
                            'A': PatternFill(start_color="FFC107", end_color="FFC107", 
                                           fill_type="solid"),
                            'B': PatternFill(start_color="FF9800", end_color="FF9800", 
                                           fill_type="solid"),
                            'C': PatternFill(start_color="FF5722", end_color="FF5722", 
                                           fill_type="solid"),
                            'D': PatternFill(start_color="F44336", end_color="F44336", 
                                           fill_type="solid")
                        }
                        
                        for row in range(2, len(self.result_df) + 2):
                            cell = worksheet.cell(row=row, column=asil_col)
                            if cell.value in asil_fills:
                                cell.fill = asil_fills[cell.value]
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.alignment = Alignment(horizontal="center")
                    
                    # Add borders
                    thin_border = Border(
                        left=Side(style='thin', color='CCCCCC'),
                        right=Side(style='thin', color='CCCCCC'),
                        top=Side(style='thin', color='CCCCCC'),
                        bottom=Side(style='thin', color='CCCCCC')
                    )
                    
                    for row in worksheet.iter_rows(min_row=1, max_row=len(self.result_df)+1):
                        for cell in row:
                            cell.border = thin_border
                
                self.log(f"Results exported to {Path(file_path).name}")
                self.status_label.setText("Exported")
                QMessageBox.information(self, "Success", f"Results exported successfully.")
                
            except Exception as e:
                QMessageBox.critical(self, "Export Error", f"Failed to export results: {str(e)}")
                self.log(f"ERROR: Export failed - {str(e)}")