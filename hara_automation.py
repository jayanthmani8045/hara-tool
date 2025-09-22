"""
HARA Automation Tool - Professional Edition with Fuzzy Matching Controls
Automated ASIL determination from HARA Excel sheets with embedded lookup table.
Version: 1.0.0
"""

import sys
import os
from pathlib import Path
import pandas as pd
from fuzzywuzzy import fuzz
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QComboBox, QTableWidget,
    QTableWidgetItem, QTextEdit, QGroupBox, QProgressBar,
    QMessageBox, QSplitter, QHeaderView, QTabWidget, QFrame,
    QCheckBox, QSlider, QDialog, QSpinBox, QFormLayout, QDialogButtonBox
)
from PySide6.QtCore import Qt, QThread, Signal, QTimer
from PySide6.QtGui import QFont, QPalette, QColor, QIcon, QPixmap, QPainter


# Embedded ASIL Determination Table
ASIL_DETERMINATION_TABLE = {
    ('S1', 'E1'): {'C0': 'QM', 'C1': 'QM', 'C2': 'QM', 'C3': 'QM'},
    ('S1', 'E2'): {'C0': 'QM', 'C1': 'QM', 'C2': 'QM', 'C3': 'QM'},
    ('S1', 'E3'): {'C0': 'QM', 'C1': 'QM', 'C2': 'QM', 'C3': 'A'},
    ('S1', 'E4'): {'C0': 'QM', 'C1': 'QM', 'C2': 'A', 'C3': 'B'},
    ('S2', 'E1'): {'C0': 'QM', 'C1': 'QM', 'C2': 'QM', 'C3': 'QM'},
    ('S2', 'E2'): {'C0': 'QM', 'C1': 'QM', 'C2': 'QM', 'C3': 'A'},
    ('S2', 'E3'): {'C0': 'QM', 'C1': 'QM', 'C2': 'A', 'C3': 'B'},
    ('S2', 'E4'): {'C0': 'A', 'C1': 'A', 'C2': 'B', 'C3': 'C'},
    ('S3', 'E1'): {'C0': 'QM', 'C1': 'QM', 'C2': 'A', 'C3': 'B'},
    ('S3', 'E2'): {'C0': 'QM', 'C1': 'QM', 'C2': 'B', 'C3': 'C'},
    ('S3', 'E3'): {'C0': 'A', 'C1': 'A', 'C2': 'B', 'C3': 'C'},
    ('S3', 'E4'): {'C0': 'B', 'C1': 'B', 'C2': 'C', 'C3': 'D'},
}


class AdvancedMatchingDialog(QDialog):
    """Dialog for advanced matching settings"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Advanced Matching Settings")
        self.setModal(True)
        self.setStyleSheet(parent.styleSheet() if parent else "")
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Case sensitivity
        self.case_sensitive = QCheckBox("Case Sensitive Matching")
        layout.addWidget(self.case_sensitive)
        
        # Strip whitespace
        self.strip_whitespace = QCheckBox("Strip Extra Whitespace")
        self.strip_whitespace.setChecked(True)
        layout.addWidget(self.strip_whitespace)
        
        # Ignore special characters
        self.ignore_special = QCheckBox("Ignore Special Characters")
        layout.addWidget(self.ignore_special)
        
        # Weight settings for combined matching
        weight_group = QGroupBox("Match Weights")
        weight_layout = QFormLayout()
        
        self.os_weight = QSpinBox()
        self.os_weight.setRange(0, 100)
        self.os_weight.setValue(70)
        self.os_weight.setSuffix("%")
        weight_layout.addRow("Operating Scenario Weight:", self.os_weight)
        
        self.hazard_weight = QSpinBox()
        self.hazard_weight.setRange(0, 100)
        self.hazard_weight.setValue(30)
        self.hazard_weight.setSuffix("%")
        weight_layout.addRow("Hazard Weight:", self.hazard_weight)
        
        weight_group.setLayout(weight_layout)
        layout.addWidget(weight_group)
        
        # Buttons
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
        self.setLayout(layout)


class ExcelProcessor(QThread):
    """Background thread for processing Excel files"""
    progress = Signal(int)
    log = Signal(str)
    finished = Signal(pd.DataFrame)
    error = Signal(str)
    
    def __init__(self, user_file, os_sheet, ra_sheet, fuzzy_enabled=True,
                 fuzzy_threshold=80, fuzzy_algorithm='Ratio (Default)',
                 case_sensitive=False, strip_whitespace=True,
                 os_weight=70, hazard_weight=30):
        super().__init__()
        self.user_file = user_file
        self.os_sheet = os_sheet
        self.ra_sheet = ra_sheet
        self.fuzzy_enabled = fuzzy_enabled
        self.fuzzy_threshold = fuzzy_threshold
        self.fuzzy_algorithm = fuzzy_algorithm
        self.case_sensitive = case_sensitive
        self.strip_whitespace = strip_whitespace
        self.os_weight = os_weight / 100.0
        self.hazard_weight = hazard_weight / 100.0
        self.result_df = None
        
    def run(self):
        """Main processing logic"""
        try:
            # Step 1: Load user's Excel sheets
            self.log.emit(f"Loading Operating Scenario sheet: {self.os_sheet}")
            self.progress.emit(20)
            os_df = pd.read_excel(self.user_file, sheet_name=self.os_sheet)
            
            self.log.emit(f"Loading Risk Assessment sheet: {self.ra_sheet}")
            self.progress.emit(30)
            ra_df = pd.read_excel(self.user_file, sheet_name=self.ra_sheet)
            
            # Step 2: Process Operating Scenarios
            self.log.emit("Processing Operating Scenarios...")
            self.progress.emit(40)
            processed_df = self.process_operating_scenarios(os_df, ra_df)
            
            # Step 3: Match with Risk Assessment
            self.log.emit(f"Matching scenarios (Fuzzy: {self.fuzzy_enabled}, Threshold: {self.fuzzy_threshold}%)")
            self.progress.emit(60)
            matched_df = self.match_risk_assessment(processed_df, ra_df)
            
            # Step 4: Determine ASIL values
            self.log.emit("Determining ASIL values...")
            self.progress.emit(80)
            final_df = self.determine_asil(matched_df)
            
            self.log.emit("Processing completed successfully")
            self.progress.emit(100)
            self.finished.emit(final_df)
            
        except Exception as e:
            self.error.emit(f"Error during processing: {str(e)}")
    
    def prepare_string(self, text):
        """Prepare string for comparison based on settings"""
        if pd.isna(text):
            return ""
        text = str(text)
        if self.strip_whitespace:
            text = ' '.join(text.split())  # Remove extra whitespace
        if not self.case_sensitive:
            text = text.lower()
        return text.strip()
    
    def get_fuzzy_score(self, str1, str2):
        """Calculate fuzzy score based on selected algorithm"""
        str1 = self.prepare_string(str1)
        str2 = self.prepare_string(str2)
        
        if not str1 or not str2:
            return 0
        
        if self.fuzzy_algorithm == 'Partial Ratio':
            return fuzz.partial_ratio(str1, str2)
        elif self.fuzzy_algorithm == 'Token Sort Ratio':
            return fuzz.token_sort_ratio(str1, str2)
        elif self.fuzzy_algorithm == 'Token Set Ratio':
            return fuzz.token_set_ratio(str1, str2)
        else:  # Default: Ratio
            return fuzz.ratio(str1, str2)
    
    def find_column_case_insensitive(self, df, column_name):
        """Find column name case-insensitively"""
        for col in df.columns:
            if str(col).strip().lower() == column_name.lower():
                return col
        return None
    
    def process_operating_scenarios(self, os_df, ra_df):
        """Extract and process Operating Scenario data"""
        results = []
        
        # Find columns case-insensitively
        os_col = self.find_column_case_insensitive(os_df, 'operating scenario')
        e_col = self.find_column_case_insensitive(os_df, 'e')
        
        # Find hazard columns (multiple possible)
        hazard_cols = [col for col in os_df.columns 
                      if 'hazard' in str(col).lower() and col != os_col]
        
        if not os_col:
            raise ValueError("Operating Scenario column not found")
        
        self.log.emit(f"Found columns: Operating Scenario, E, {len(hazard_cols)} hazard columns")
        
        # Process each row
        for idx, row in os_df.iterrows():
            # Skip if Operating Scenario is empty
            if pd.isna(row.get(os_col, None)) or str(row.get(os_col, '')).strip() == '':
                continue
            
            operating_scenario = str(row[os_col]).strip()
            
            # Get E value (default to 4 if empty)
            e_value = 4
            if e_col and not pd.isna(row.get(e_col, None)):
                try:
                    e_value = int(row[e_col])
                except:
                    e_value = 4
            
            # Get hazards from all hazard columns
            hazards = []
            for h_col in hazard_cols:
                if not pd.isna(row.get(h_col, None)) and str(row.get(h_col, '')).strip():
                    hazards.append(str(row[h_col]).strip())
            
            # If no hazard found, we'll match from Risk Assessment later
            hazard = hazards[0] if hazards else None
            
            results.append({
                'Operating Scenario': operating_scenario,
                'E': e_value,
                'Hazard': hazard,
                'Original_Index': idx
            })
        
        self.log.emit(f"Processed {len(results)} operating scenarios")
        return pd.DataFrame(results)
    
    def match_risk_assessment(self, processed_df, ra_df):
        """Match processed scenarios with Risk Assessment data"""
        
        # Find columns in Risk Assessment
        ra_os_col = self.find_column_case_insensitive(ra_df, 'operating scenario')
        ra_hazard_col = self.find_column_case_insensitive(ra_df, 'hazard')
        ra_he_col = self.find_column_case_insensitive(ra_df, 'hazardous event')
        ra_details_col = self.find_column_case_insensitive(ra_df, 'details of hazardous event')
        ra_people_col = self.find_column_case_insensitive(ra_df, 'people at risk')
        ra_delta_col = self.find_column_case_insensitive(ra_df, 'δv') or self.find_column_case_insensitive(ra_df, 'deltav') or self.find_column_case_insensitive(ra_df, 'Δv')
        ra_s_col = self.find_column_case_insensitive(ra_df, 's')
        ra_s_rational_col = self.find_column_case_insensitive(ra_df, 'severity rational')
        ra_c_col = self.find_column_case_insensitive(ra_df, 'c')
        ra_c_rational_col = self.find_column_case_insensitive(ra_df, 'controllability rational')
        
        matched_results = []
        exact_matches = 0
        fuzzy_matches = 0
        no_matches = 0
        
        for _, row in processed_df.iterrows():
            os_text = row['Operating Scenario']
            hazard = row['Hazard']
            e_value = row['E']
            
            # Try exact match first
            best_match = None
            best_score = 0
            match_type = "none"
            match_details = ""
            
            # Prepare text for comparison
            os_text_prepared = self.prepare_string(os_text)
            
            # Exact matching phase
            for _, ra_row in ra_df.iterrows():
                if pd.isna(ra_row.get(ra_os_col, None)):
                    continue
                    
                ra_os_text = self.prepare_string(ra_row[ra_os_col])
                
                # Exact match
                if os_text_prepared == ra_os_text:
                    # If hazard specified, check it matches
                    if hazard and ra_hazard_col:
                        hazard_prepared = self.prepare_string(hazard)
                        if pd.notna(ra_row.get(ra_hazard_col)):
                            ra_hazard = self.prepare_string(ra_row[ra_hazard_col])
                            if hazard_prepared == ra_hazard:
                                best_match = ra_row
                                best_score = 100
                                match_type = "exact"
                                match_details = "Exact (OS+Hazard)"
                                break
                    else:
                        best_match = ra_row
                        best_score = 100
                        match_type = "exact"
                        match_details = "Exact (OS)"
                        break
            
            # Fuzzy matching phase (if enabled and no exact match)
            if best_score < 100 and self.fuzzy_enabled:
                for _, ra_row in ra_df.iterrows():
                    if pd.isna(ra_row.get(ra_os_col, None)):
                        continue
                    
                    ra_os_text = str(ra_row[ra_os_col]).strip()
                    
                    # Calculate fuzzy match score using selected algorithm
                    os_score = self.get_fuzzy_score(os_text, ra_os_text)
                    
                    # Calculate combined score with hazard if available
                    if hazard and ra_hazard_col and pd.notna(ra_row.get(ra_hazard_col)):
                        hazard_score = self.get_fuzzy_score(hazard, str(ra_row[ra_hazard_col]))
                        # Weighted average based on user settings
                        score = (os_score * self.os_weight) + (hazard_score * self.hazard_weight)
                    else:
                        score = os_score
                    
                    # Check if score meets threshold
                    if score > best_score and score >= self.fuzzy_threshold:
                        best_match = ra_row
                        best_score = score
                        match_type = "fuzzy"
                        algorithm_name = self.fuzzy_algorithm.split(' ')[0]
                        match_details = f"Fuzzy-{algorithm_name} ({score:.0f}%)"
            
            # Count match types
            if match_type == "exact":
                exact_matches += 1
            elif match_type == "fuzzy":
                fuzzy_matches += 1
            else:
                no_matches += 1
            
            # Prepare matched row with match information
            matched_row = {
                'Operating Scenario': os_text,
                'E': e_value,
                'Hazard': hazard if hazard else (best_match[ra_hazard_col] if best_match is not None and ra_hazard_col else ''),
                'Match_Type': match_details if match_details else 'No match',
                'Match_Score': f"{best_score:.0f}%" if best_score > 0 else "0%"
            }
            
            # Add fields from Risk Assessment if match found
            if best_match is not None:
                if ra_he_col:
                    matched_row['Hazardous Event'] = best_match.get(ra_he_col, '')
                if ra_details_col:
                    matched_row['Details of Hazardous event'] = best_match.get(ra_details_col, '')
                if ra_people_col:
                    matched_row['people at risk'] = best_match.get(ra_people_col, '')
                if ra_delta_col:
                    matched_row['Δv'] = best_match.get(ra_delta_col, '')
                
                # Only set S if valid value found (no default)
                if ra_s_col and pd.notna(best_match.get(ra_s_col)):
                    try:
                        s_val = int(best_match.get(ra_s_col))
                        if 0 <= s_val <= 3:  # Validate S range
                            matched_row['S'] = s_val
                        else:
                            matched_row['S'] = ''
                    except:
                        matched_row['S'] = ''
                else:
                    matched_row['S'] = ''
                
                if ra_s_rational_col:
                    matched_row['Severity Rational'] = best_match.get(ra_s_rational_col, '')
                
                # Only set C if valid value found (no default)
                if ra_c_col and pd.notna(best_match.get(ra_c_col)):
                    try:
                        c_val = int(best_match.get(ra_c_col))
                        if 0 <= c_val <= 3:  # Validate C range
                            matched_row['C'] = c_val
                        else:
                            matched_row['C'] = ''
                    except:
                        matched_row['C'] = ''
                else:
                    matched_row['C'] = ''
                    
                if ra_c_rational_col:
                    matched_row['Controllability Rational'] = best_match.get(ra_c_rational_col, '')
            else:
                # No match found - leave S and C empty (no defaults)
                matched_row.update({
                    'Hazardous Event': '',
                    'Details of Hazardous event': '',
                    'people at risk': '',
                    'Δv': '',
                    'S': '',
                    'Severity Rational': '',
                    'C': '',
                    'Controllability Rational': ''
                })
            
            matched_results.append(matched_row)
        
        algorithm_info = f" ({self.fuzzy_algorithm.split(' ')[0]})" if self.fuzzy_enabled else ""
        self.log.emit(f"Matching complete: {exact_matches} exact, {fuzzy_matches} fuzzy{algorithm_info}, {no_matches} unmatched")
        return pd.DataFrame(matched_results)
    
    def determine_asil(self, matched_df):
        """Determine ASIL values based on E, S, C using embedded lookup table"""
        
        def get_asil_value(e, s, c):
            """Lookup ASIL value from embedded determination table"""
            # Only calculate ASIL if S and C are valid values
            if pd.isna(s) or s == '' or pd.isna(c) or c == '':
                return ''  # Return empty if S or C is missing
            
            try:
                # Convert values to valid ranges
                e_val = min(max(int(e), 1), 4)
                s_val = int(s)
                c_val = int(c)
                
                # Validate ranges
                if not (0 <= s_val <= 3) or not (0 <= c_val <= 3):
                    return ''
                
                # Create lookup keys
                s_key = f"S{s_val}"
                e_key = f"E{e_val}"
                c_key = f"C{c_val}"
                
                # Lookup in embedded table
                if (s_key, e_key) in ASIL_DETERMINATION_TABLE:
                    return ASIL_DETERMINATION_TABLE[(s_key, e_key)].get(c_key, '')
            except:
                return ''
            
            return ''  # Return empty if lookup fails
        
        # Add ASIL column
        matched_df['ASIL'] = matched_df.apply(
            lambda row: get_asil_value(row['E'], row['S'], row['C']), 
            axis=1
        )
        
        # Count ASIL distribution
        asil_values = matched_df['ASIL'][matched_df['ASIL'] != '']
        if len(asil_values) > 0:
            asil_counts = asil_values.value_counts()
            self.log.emit(f"ASIL determination: {', '.join([f'{val}={count}' for val, count in asil_counts.items()])}")
        else:
            self.log.emit("ASIL determination: No valid S/C values found for ASIL calculation")
        
        return matched_df


class HARAMainWindow(QMainWindow):
    """Main application window with professional black and light blue theme"""
    
    def __init__(self):
        super().__init__()
        self.user_file = None
        self.result_df = None
        self.advanced_settings = {
            'case_sensitive': False,
            'strip_whitespace': True,
            'os_weight': 70,
            'hazard_weight': 30
        }
        
        # Set window icon
        self.set_window_icon()
        self.init_ui()
    
    def set_window_icon(self):
        """Set application icon"""
        # Only use icon.ico file
        if os.path.exists('icon.ico'):
            icon = QIcon('icon.ico')
            self.setWindowIcon(icon)
    
    def init_ui(self):
        """Initialize the user interface with professional theme"""
        self.setWindowTitle("HARA Automation Tool - Professional Edition")
        self.setGeometry(100, 100, 1500, 900)
        
        # Apply professional black and light blue theme
        self.setStyleSheet("""
            QMainWindow {
                background-color: #0a0a0a;
                color: #e8e8e8;
            }
            QWidget {
                background-color: #0a0a0a;
                color: #e8e8e8;
                font-family: 'Segoe UI', 'Arial', sans-serif;
            }
            QGroupBox {
                font-weight: 600;
                font-size: 11pt;
                border: 1px solid #2c4f7c;
                border-radius: 6px;
                margin-top: 12px;
                padding-top: 15px;
                background-color: #0f0f0f;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 10px 0 10px;
                color: #5eb3f6;
                background-color: #0f0f0f;
            }
            QPushButton {
                background-color: #1e3a5f;
                color: #ffffff;
                border: 1px solid #2c4f7c;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: 500;
                font-size: 10pt;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #2c4f7c;
                border: 1px solid #5eb3f6;
            }
            QPushButton:pressed {
                background-color: #1a2940;
            }
            QPushButton:disabled {
                background-color: #141414;
                color: #4a4a4a;
                border: 1px solid #2a2a2a;
            }
            QLabel {
                color: #c8d4e0;
                font-size: 10pt;
            }
            QComboBox {
                background-color: #141414;
                border: 1px solid #2c4f7c;
                border-radius: 4px;
                padding: 6px;
                min-width: 150px;
                color: #e8e8e8;
                font-size: 10pt;
            }
            QComboBox:hover {
                border: 1px solid #5eb3f6;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 4px solid #5eb3f6;
                margin-right: 5px;
            }
            QComboBox QAbstractItemView {
                background-color: #141414;
                border: 1px solid #2c4f7c;
                selection-background-color: #2c4f7c;
                selection-color: #ffffff;
            }
            QCheckBox {
                color: #e8e8e8;
                font-size: 10pt;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border: 1px solid #2c4f7c;
                border-radius: 3px;
                background-color: #141414;
            }
            QCheckBox::indicator:checked {
                background-color: #5eb3f6;
                border: 1px solid #5eb3f6;
            }
            QSlider::groove:horizontal {
                height: 6px;
                background: #2c4f7c;
                border-radius: 3px;
            }
            QSlider::handle:horizontal {
                background: #5eb3f6;
                border: 1px solid #5eb3f6;
                width: 18px;
                height: 18px;
                margin: -6px 0;
                border-radius: 9px;
            }
            QSlider::handle:horizontal:hover {
                background: #7fc4f8;
            }
            QTableWidget {
                background-color: #0f0f0f;
                alternate-background-color: #141414;
                gridline-color: #1e3a5f;
                border: 1px solid #2c4f7c;
                border-radius: 4px;
                color: #e8e8e8;
                font-size: 9pt;
            }
            QTableWidget::item {
                padding: 5px;
                border: none;
            }
            QTableWidget::item:selected {
                background-color: #2c4f7c;
                color: #ffffff;
            }
            QHeaderView::section {
                background-color: #1e3a5f;
                color: #ffffff;
                font-weight: 600;
                padding: 8px;
                border: none;
                border-right: 1px solid #0a0a0a;
                border-bottom: 1px solid #2c4f7c;
            }
            QTextEdit {
                background-color: #050505;
                color: #5eb3f6;
                font-family: 'Consolas', 'Courier New', monospace;
                font-size: 9pt;
                border: 1px solid #2c4f7c;
                border-radius: 4px;
                padding: 8px;
            }
            QProgressBar {
                border: 1px solid #2c4f7c;
                border-radius: 4px;
                text-align: center;
                background-color: #141414;
                color: #e8e8e8;
                font-weight: 500;
                height: 22px;
            }
            QProgressBar::chunk {
                background-color: #5eb3f6;
                border-radius: 3px;
            }
            QTabWidget::pane {
                border: 1px solid #2c4f7c;
                background-color: #0f0f0f;
                border-radius: 4px;
            }
            QTabBar::tab {
                background-color: #141414;
                color: #808080;
                padding: 8px 16px;
                margin-right: 2px;
                border: 1px solid #1e3a5f;
                border-bottom: none;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background-color: #0f0f0f;
                color: #5eb3f6;
                font-weight: 500;
                border-color: #2c4f7c;
            }
            QTabBar::tab:hover:!selected {
                background-color: #1a1a1a;
                color: #c8d4e0;
            }
            QMessageBox {
                background-color: #141414;
                color: #e8e8e8;
            }
            QMessageBox QPushButton {
                min-width: 70px;
            }
            QFrame#separator {
                background-color: #2c4f7c;
                height: 1px;
                border: none;
            }
        """)
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Title Section
        title_frame = QFrame()
        title_layout = QVBoxLayout(title_frame)
        
        title_label = QLabel("HARA AUTOMATION TOOL")
        title_label.setStyleSheet("""
            font-size: 20pt; 
            font-weight: 300; 
            color: #5eb3f6; 
            padding: 15px;
            letter-spacing: 2px;
        """)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_layout.addWidget(title_label)
        
        subtitle_label = QLabel("Automated Safety Integrity Level Determination")
        subtitle_label.setStyleSheet("""
            font-size: 10pt; 
            color: #808080; 
            padding-bottom: 10px;
        """)
        subtitle_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_layout.addWidget(subtitle_label)
        
        # Separator line
        separator = QFrame()
        separator.setObjectName("separator")
        separator.setFixedHeight(1)
        title_layout.addWidget(separator)
        
        main_layout.addWidget(title_frame)
        
        # File Selection Section
        file_group = QGroupBox("FILE SELECTION")
        file_layout = QVBoxLayout()
        
        # User file selection
        user_file_layout = QHBoxLayout()
        file_label_title = QLabel("Excel File:")
        file_label_title.setFixedWidth(100)
        user_file_layout.addWidget(file_label_title)
        
        self.file_label = QLabel("No file selected")
        self.file_label.setStyleSheet("""
            padding: 8px; 
            background-color: #141414; 
            border: 1px solid #2c4f7c;
            border-radius: 4px;
            color: #808080;
        """)
        user_file_layout.addWidget(self.file_label, 1)
        
        self.browse_btn = QPushButton("Browse")
        self.browse_btn.clicked.connect(self.browse_file)
        user_file_layout.addWidget(self.browse_btn)
        file_layout.addLayout(user_file_layout)
        
        # Sheet selection
        sheet_layout = QHBoxLayout()
        
        os_label = QLabel("Operating Scenario Sheet:")
        os_label.setFixedWidth(180)
        sheet_layout.addWidget(os_label)
        self.os_combo = QComboBox()
        self.os_combo.setEnabled(False)
        sheet_layout.addWidget(self.os_combo, 1)
        
        sheet_layout.addSpacing(30)
        
        ra_label = QLabel("Risk Assessment Sheet:")
        ra_label.setFixedWidth(180)
        sheet_layout.addWidget(ra_label)
        self.ra_combo = QComboBox()
        self.ra_combo.setEnabled(False)
        sheet_layout.addWidget(self.ra_combo, 1)
        
        file_layout.addLayout(sheet_layout)
        file_group.setLayout(file_layout)
        main_layout.addWidget(file_group)
        
        # Processing Section with Fuzzy Matching Controls
        process_group = QGroupBox("PROCESSING & MATCHING SETTINGS")
        process_layout = QVBoxLayout()
        
        # Fuzzy Matching Control Section
        fuzzy_layout = QHBoxLayout()
        
        # Fuzzy matching checkbox
        self.fuzzy_enabled = QCheckBox("Enable Fuzzy Matching")
        self.fuzzy_enabled.setChecked(True)
        self.fuzzy_enabled.setStyleSheet("color: #5eb3f6; font-weight: 500;")
        self.fuzzy_enabled.toggled.connect(self.on_fuzzy_toggle)
        fuzzy_layout.addWidget(self.fuzzy_enabled)
        
        # Threshold slider
        fuzzy_layout.addWidget(QLabel("Threshold:"))
        self.fuzzy_slider = QSlider(Qt.Orientation.Horizontal)
        self.fuzzy_slider.setMinimum(50)
        self.fuzzy_slider.setMaximum(100)
        self.fuzzy_slider.setValue(80)  # Default 80%
        self.fuzzy_slider.setTickPosition(QSlider.TickPosition.TicksBelow)
        self.fuzzy_slider.setTickInterval(10)
        self.fuzzy_slider.setFixedWidth(200)
        self.fuzzy_slider.valueChanged.connect(self.on_threshold_change)
        fuzzy_layout.addWidget(self.fuzzy_slider)
        
        # Threshold value label
        self.threshold_label = QLabel("80%")
        self.threshold_label.setStyleSheet("color: #5eb3f6; font-weight: bold; min-width: 40px;")
        fuzzy_layout.addWidget(self.threshold_label)
        
        # Matching algorithm selection
        fuzzy_layout.addWidget(QLabel("Algorithm:"))
        self.algorithm_combo = QComboBox()
        self.algorithm_combo.addItems([
            "Ratio (Default)",
            "Partial Ratio",
            "Token Sort Ratio",
            "Token Set Ratio"
        ])
        self.algorithm_combo.setToolTip(
            "Ratio: Simple string similarity\n"
            "Partial: Matches substring\n"
            "Token Sort: Ignores word order\n"
            "Token Set: Ignores duplicates and order"
        )
        fuzzy_layout.addWidget(self.algorithm_combo)
        
        # Advanced settings button
        self.advanced_btn = QPushButton("Advanced")
        self.advanced_btn.clicked.connect(self.show_advanced_settings)
        fuzzy_layout.addWidget(self.advanced_btn)
        
        fuzzy_layout.addStretch()
        process_layout.addLayout(fuzzy_layout)
        
        # Process button and progress
        process_btn_layout = QHBoxLayout()
        self.process_btn = QPushButton("Process HARA")
        self.process_btn.setEnabled(False)
        self.process_btn.setStyleSheet("""
            QPushButton:enabled {
                background-color: #2c4f7c;
                font-weight: 600;
            }
            QPushButton:enabled:hover {
                background-color: #5eb3f6;
                color: #0a0a0a;
            }
        """)
        self.process_btn.clicked.connect(self.process_data)
        process_btn_layout.addWidget(self.process_btn)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setTextVisible(True)
        process_btn_layout.addWidget(self.progress_bar, 1)
        
        # Status label
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("color: #5eb3f6; font-weight: 500;")
        process_btn_layout.addWidget(self.status_label)
        
        process_layout.addLayout(process_btn_layout)
        process_group.setLayout(process_layout)
        main_layout.addWidget(process_group)
        
        # Results Section (Tabs)
        self.tabs = QTabWidget()
        
        # Results Table Tab
        self.result_table = QTableWidget()
        self.result_table.setAlternatingRowColors(True)
        self.tabs.addTab(self.result_table, "Results")
        
        # Log Tab
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.tabs.addTab(self.log_text, "Processing Log")
        
        main_layout.addWidget(self.tabs, 1)
        
        # Save Section
        save_group = QGroupBox("EXPORT OPTIONS")
        save_layout = QHBoxLayout()
        
        self.save_btn = QPushButton("Save to Excel")
        self.save_btn.setToolTip("Add RESULT sheet to original file")
        self.save_btn.setEnabled(False)
        self.save_btn.clicked.connect(self.save_results)
        save_layout.addWidget(self.save_btn)
        
        self.export_btn = QPushButton("Export New File")
        self.export_btn.setToolTip("Export results to a new Excel file")
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_results)
        save_layout.addWidget(self.export_btn)
        
        save_layout.addStretch()
        
        # Info Label
        info_label = QLabel("ASIL Determination: Embedded | Valid Ranges: S(0-3) C(0-3) E(1-4)")
        info_label.setStyleSheet("color: #4a5a6a; font-size: 9pt;")
        save_layout.addWidget(info_label)
        
        save_group.setLayout(save_layout)
        main_layout.addWidget(save_group)
        
        # Initial log message
        self.log("System initialized. ASIL determination table embedded.")
        self.log("Fuzzy matching enabled with 80% threshold (Ratio algorithm).")
    
    def on_fuzzy_toggle(self, checked):
        """Handle fuzzy matching toggle"""
        self.fuzzy_slider.setEnabled(checked)
        self.algorithm_combo.setEnabled(checked)
        if checked:
            self.log(f"Fuzzy matching enabled (threshold: {self.fuzzy_slider.value()}%)")
        else:
            self.log("Fuzzy matching disabled - exact matches only")
    
    def on_threshold_change(self, value):
        """Handle threshold slider change"""
        self.threshold_label.setText(f"{value}%")
    
    def show_advanced_settings(self):
        """Show advanced matching settings dialog"""
        dialog = AdvancedMatchingDialog(self)
        
        # Load current settings
        dialog.case_sensitive.setChecked(self.advanced_settings['case_sensitive'])
        dialog.strip_whitespace.setChecked(self.advanced_settings['strip_whitespace'])
        dialog.os_weight.setValue(self.advanced_settings['os_weight'])
        dialog.hazard_weight.setValue(self.advanced_settings['hazard_weight'])
        
        if dialog.exec():
            # Save settings
            self.advanced_settings['case_sensitive'] = dialog.case_sensitive.isChecked()
            self.advanced_settings['strip_whitespace'] = dialog.strip_whitespace.isChecked()
            self.advanced_settings['os_weight'] = dialog.os_weight.value()
            self.advanced_settings['hazard_weight'] = dialog.hazard_weight.value()
            
            self.log(f"Advanced settings updated: Case sensitive={self.advanced_settings['case_sensitive']}, "
                    f"Strip whitespace={self.advanced_settings['strip_whitespace']}, "
                    f"Weights: OS={self.advanced_settings['os_weight']}%, Hazard={self.advanced_settings['hazard_weight']}%")
    
    def browse_file(self):
        """Browse for user's Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            self.user_file = file_path
            self.file_label.setText(Path(file_path).name)
            self.file_label.setStyleSheet("""
                padding: 8px; 
                background-color: #141414; 
                border: 1px solid #5eb3f6;
                border-radius: 4px;
                color: #5eb3f6;
                font-weight: 500;
            """)
            self.load_sheets()
            
    def load_sheets(self):
        """Load sheet names from Excel file"""
        try:
            xl_file = pd.ExcelFile(str(self.user_file))
            sheets = xl_file.sheet_names
            
            # Filter out HAZOP sheet if it exists
            filtered_sheets = [str(sheet) for sheet in sheets if 'hazop' not in str(sheet).lower()]
            
            self.os_combo.clear()
            self.os_combo.addItems((filtered_sheets))
            self.os_combo.setEnabled(True)
            
            self.ra_combo.clear()
            self.ra_combo.addItems((filtered_sheets))
            self.ra_combo.setEnabled(True)
            
            # Try to auto-select sheets based on names
            for i, sheet in enumerate(filtered_sheets):
                sheet_lower = sheet.lower()
                if 'operating' in sheet_lower and 'scenario' in sheet_lower:
                    self.os_combo.setCurrentIndex(i)
                elif 'risk' in sheet_lower and 'assessment' in sheet_lower:
                    self.ra_combo.setCurrentIndex(i)
            
            self.process_btn.setEnabled(True)
            self.log(f"File loaded: {len(filtered_sheets)} sheets available")
            self.status_label.setText("Ready")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load Excel file: {str(e)}")
            self.log(f"ERROR: Failed to load file - {str(e)}")
    
    def process_data(self):
        """Start processing in background thread"""
        if not self.user_file:
            QMessageBox.warning(self, "Warning", "Please select an Excel file first")
            return
            
        os_sheet = self.os_combo.currentText()
        ra_sheet = self.ra_combo.currentText()
        
        if os_sheet == ra_sheet:
            QMessageBox.warning(self, "Warning", "Please select different sheets for Operating Scenario and Risk Assessment")
            return
        
        # Get fuzzy matching settings
        fuzzy_enabled = self.fuzzy_enabled.isChecked()
        fuzzy_threshold = self.fuzzy_slider.value()
        fuzzy_algorithm = self.algorithm_combo.currentText()
        
        # Disable controls during processing
        self.process_btn.setEnabled(False)
        self.save_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("Processing...")
        
        # Clear previous results
        self.result_table.clear()
        self.tabs.setCurrentIndex(1)  # Switch to log tab
        
        # Create and start processor thread with all settings
        self.processor = ExcelProcessor(
            self.user_file,
            os_sheet,
            ra_sheet,
            fuzzy_enabled,
            fuzzy_threshold,
            fuzzy_algorithm,
            self.advanced_settings['case_sensitive'],
            self.advanced_settings['strip_whitespace'],
            self.advanced_settings['os_weight'],
            self.advanced_settings['hazard_weight']
        )
        
        self.processor.progress.connect(self.update_progress)
        self.processor.log.connect(self.log)
        self.processor.finished.connect(self.on_process_complete)
        self.processor.error.connect(self.on_process_error)
        
        self.log(f"Processing started: {os_sheet} -> {ra_sheet}")
        self.processor.start()
    
    def update_progress(self, value):
        """Update progress bar"""
        self.progress_bar.setValue(value)
    
    def log(self, message):
        """Add message to log with timestamp"""
        timestamp = pd.Timestamp.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{timestamp}] {message}")
    
    def on_process_complete(self, result_df):
        """Handle successful processing"""
        self.result_df = result_df
        
        # Display results in table
        self.display_results(result_df)
        
        # Enable save buttons
        self.save_btn.setEnabled(True)
        self.export_btn.setEnabled(True)
        
        # Re-enable controls
        self.process_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        
        # Count ASIL statistics
        asil_count = len(result_df[result_df['ASIL'] != ''])
        total_count = len(result_df)
        self.status_label.setText(f"Complete: {total_count} rows processed, {asil_count} with ASIL")
        
        # Switch to results tab
        self.tabs.setCurrentIndex(0)
        
        # Show summary
        if asil_count > 0:
            asil_summary = result_df[result_df['ASIL'] != '']['ASIL'].value_counts().to_dict()
            summary_text = ', '.join([f"{k}: {v}" for k, v in sorted(asil_summary.items())])
            message = f"Processing completed successfully.\n\nTotal Scenarios: {total_count}\nASIL Determined: {asil_count}\n\nDistribution:\n{summary_text}"
        else:
            message = f"Processing completed.\n\nTotal Scenarios: {total_count}\nNo ASIL values determined (S/C values not found)"
        
        QMessageBox.information(self, "Processing Complete", message)
    
    def on_process_error(self, error_msg):
        """Handle processing error"""
        self.log(f"ERROR: {error_msg}")
        self.status_label.setText("Error")
        QMessageBox.critical(self, "Processing Error", error_msg)
        
        # Re-enable controls
        self.process_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
    
    def display_results(self, df):
        """Display DataFrame in table widget with color coding"""
        self.result_table.setRowCount(len(df))
        self.result_table.setColumnCount(len(df.columns))
        self.result_table.setHorizontalHeaderLabels(df.columns.tolist())
        
        # ASIL color mapping (professional colors)
        asil_colors = {
            'QM': QColor(46, 125, 50),       # Green
            'A': QColor(255, 193, 7),        # Amber
            'B': QColor(255, 152, 0),        # Orange
            'C': QColor(255, 87, 34),        # Deep Orange
            'D': QColor(244, 67, 54)         # Red
        }
        
        for row in range(len(df)):
            for col in range(len(df.columns)):
                value = df.iloc[row, col]
                if pd.isna(value):
                    value = ""
                item = QTableWidgetItem(str(value))
                
                # Color code ASIL column
                if df.columns[col] == 'ASIL' and value in asil_colors:
                    item.setBackground(asil_colors[value])
                    item.setForeground(QColor(255, 255, 255))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                
                # Highlight match type column
                if df.columns[col] == 'Match_Type':
                    if 'Exact' in str(value):
                        item.setForeground(QColor(46, 125, 50))  # Green for exact
                    elif 'Fuzzy' in str(value):
                        item.setForeground(QColor(94, 179, 246))  # Light blue for fuzzy
                    else:
                        item.setForeground(QColor(255, 152, 0))  # Orange for no match
                
                self.result_table.setItem(row, col, item)
        
        # Adjust column widths
        self.result_table.resizeColumnsToContents()
        header = self.result_table.horizontalHeader()
        header.setStretchLastSection(True)
    
    def save_results(self):
        """Save results to RESULT sheet in original file"""
        if self.result_df is None:
            return
        
        try:
            # Read existing Excel file
            with pd.ExcelFile(str(self.user_file)) as xl:
                sheets = {sheet: xl.parse(sheet) for sheet in xl.sheet_names}
            
            # Add or replace RESULT sheet
            sheets['RESULT'] = self.result_df
            
            # Write back to file
            with pd.ExcelWriter(str(self.user_file), engine='openpyxl') as writer:
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, sheet_name=str(sheet_name), index=False)
            
            self.log(f"Results saved to RESULT sheet")
            self.status_label.setText("Saved")
            QMessageBox.information(self, "Success", "Results saved to RESULT sheet successfully.")
            
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save results: {str(e)}")
            self.log(f"ERROR: Save failed - {str(e)}")
    
    def export_results(self):
        """Export results to a new Excel file"""
        if self.result_df is None:
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Results",
            "HARA_Results.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                # Export with formatting
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    self.result_df.to_excel(writer, sheet_name='RESULT', index=False)
                    
                    # Get the workbook and worksheet
                    workbook = writer.book
                    worksheet = writer.sheets['RESULT']
                    
                    # Apply professional formatting
                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                    
                    # Header formatting
                    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                    
                    # ASIL column formatting
                    asil_col = None
                    for idx, col in enumerate(self.result_df.columns, 1):
                        if col == 'ASIL':
                            asil_col = idx
                            break
                    
                    if asil_col:
                        asil_fills = {
                            'QM': PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),
                            'A': PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),
                            'B': PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid"),
                            'C': PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid"),
                            'D': PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
                        }
                        
                        for row in range(2, len(self.result_df) + 2):
                            cell = worksheet.cell(row=row, column=asil_col)
                            if cell.value in asil_fills:
                                cell.fill = asil_fills[cell.value]
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.alignment = Alignment(horizontal="center")
                    
                    # Add borders
                    thin_border = Border(
                        left=Side(style='thin', color='CCCCCC'),
                        right=Side(style='thin', color='CCCCCC'),
                        top=Side(style='thin', color='CCCCCC'),
                        bottom=Side(style='thin', color='CCCCCC')
                    )
                    
                    for row in worksheet.iter_rows(min_row=1, max_row=len(self.result_df)+1):
                        for cell in row:
                            cell.border = thin_border
                
                self.log(f"Results exported to {Path(file_path).name}")
                self.status_label.setText("Exported")
                QMessageBox.information(self, "Success", f"Results exported successfully.")
                
            except Exception as e:
                QMessageBox.critical(self, "Export Error", f"Failed to export results: {str(e)}")
                self.log(f"ERROR: Export failed - {str(e)}")


def main():
    """Main application entry point"""
    app = QApplication(sys.argv)
    
    # Set application metadata
    app.setApplicationName("HARA Automation Tool")
    app.setOrganizationName("Automotive Safety Systems")
    app.setApplicationDisplayName("HARA Automation")
    
    # Set application ID for Windows taskbar
    if sys.platform == 'win32':
        import ctypes
        myappid = 'AutomotiveSafety.HARAAutomation.Tool.1.0'
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
    
    # Set application icon if available
    if os.path.exists('icon.ico'):
        icon = QIcon('icon.ico')
        app.setWindowIcon(icon)
    
    # Create and show main window
    window = HARAMainWindow()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()